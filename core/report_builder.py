from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import CheckResult, DocumentRow, MatchResult
from core.paths import REPORT_DIR


MONEY_DIFF_THRESHOLD_PERCENT = Decimal("0.5")
MISSING_EXTRA_SHEET_TITLE = "Нет-Лишнее в счете"

DISCREPANCY_HEADERS = [
    "Статус",
    "Наименование",
    "Артикул заказа",
    "Артикул счета",
    "Штрихкод заказа",
    "Штрихкод счета",
    "Кол-во заказ",
    "Кол-во счет",
    "Разница по количеству",
    "Цена заказ",
    "Цена счет",
    "Сумма заказ",
    "Сумма счет",
    "Разница по сумме, ₽",
    "Разница по сумме, %",
    "Комментарий",
]

MISSING_EXTRA_HEADERS = [
    "Статус",
    "Наименование",
    "Артикул",
    "Штрихкод",
    "Кол-во заказ",
    "Кол-во счет",
    "Цена заказ",
    "Цена счет",
    "Сумма заказ",
    "Сумма счет",
    "Комментарий",
]


def build_report(result: CheckResult, output: str | Path | None = None) -> Path:
    output_path = Path(output) if output else REPORT_DIR / (
        f"doccheck_report_{datetime.now():%Y-%m-%d_%H-%M}.xlsx"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Итог"
    summary_sheet.append(["Показатель", "Значение"])
    summary_sheet.append(["Дата проверки", datetime.now().strftime("%Y-%m-%d %H:%M")])
    summary_sheet.append(["Файл заказа", result.order.path.name])
    summary_sheet.append(["Файл счета", result.invoice.path.name])
    hidden_small_money = _hidden_small_money_count(result)
    for key, value in _visible_summary(result).items():
        summary_sheet.append([key, value])
    summary_sheet.append(
        ["Скрыто мелких денежных расхождений < 0,5%", hidden_small_money]
    )

    discrepancies_sheet = workbook.create_sheet("Расхождения")
    discrepancies_sheet.append(DISCREPANCY_HEADERS)
    for row in result.rows:
        if _show_as_discrepancy(row):
            discrepancies_sheet.append(_discrepancy_values(row))

    missing_extra_sheet = workbook.create_sheet(MISSING_EXTRA_SHEET_TITLE)
    missing_extra_sheet.append(MISSING_EXTRA_HEADERS)
    for row in result.rows:
        if row.status in {"Нет в счете", "Лишнее в счете"}:
            missing_extra_sheet.append(_missing_extra_values(row))

    manual_sheet = workbook.create_sheet("Проверить вручную")
    manual_sheet.append(DISCREPANCY_HEADERS)
    for row in result.rows:
        if _is_manual(row) and not _is_hidden_small_money(row):
            manual_sheet.append(_discrepancy_values(row))

    tech = workbook.create_sheet("Тех_данные")
    tech.append(
        [
            "Источник",
            "Строка Excel",
            "item_id",
            "Наименование",
            "Штрихкод",
            "Количество",
            "Цена",
            "Сумма",
            "Ошибки",
            "Технический комментарий",
            "aggregation_group_id",
            "aggregation_status",
            "aggregation_rows_count",
            "aggregation_source_rows",
            "aggregation_key_type",
        ]
    )
    for document in (result.order, result.invoice):
        tech.append([f"Заголовок {document.source}", document.header_row, "", str(document.columns)])
        for row in document.rows:
            tech.append(
                [
                    row.source,
                    row.row_number,
                    row.item_id or "",
                    row.name,
                    row.barcode,
                    _value(row.quantity),
                    _value(row.price),
                    _value(row.amount),
                    "; ".join(row.errors),
                    "",
                    row.raw.get("aggregation_group_id", ""),
                    row.raw.get("aggregation_status", ""),
                    row.raw.get("aggregation_rows_count", ""),
                    row.raw.get("aggregation_source_rows", ""),
                    row.raw.get("aggregation_key_type", ""),
                ]
            )
    tech.append([])
    tech.append(
        [
            "Результат",
            "Статус",
            "Комментарий менеджеру",
            "Технический комментарий",
            "Заказ строка",
            "Счет строка",
            "Заказ штрихкод",
            "Счет штрихкод",
            "aggregation_group_id",
            "aggregation_status",
            "aggregation_rows_count",
            "aggregation_source_rows",
            "aggregation_key_type",
        ]
    )
    for row in result.rows:
        invoice_raw = row.invoice.raw if row.invoice else {}
        tech.append(
            [
                "Результат",
                row.status,
                _manager_comment(row),
                row.technical_comment or row.comment,
                row.order.row_number if row.order else "",
                row.invoice.row_number if row.invoice else "",
                row.order.barcode if row.order else "",
                row.invoice.barcode if row.invoice else "",
                invoice_raw.get("aggregation_group_id", ""),
                invoice_raw.get("aggregation_status", ""),
                invoice_raw.get("aggregation_rows_count", ""),
                invoice_raw.get("aggregation_source_rows", ""),
                invoice_raw.get("aggregation_key_type", ""),
            ]
        )

    for sheet in workbook.worksheets:
        _style(sheet)
    workbook.save(output_path)
    return output_path


def _discrepancy_values(result: MatchResult) -> list[object]:
    order, invoice = result.order, result.invoice
    return [
        result.status,
        _name(order, invoice),
        _attr(order, "article"),
        _attr(invoice, "article"),
        _attr(order, "barcode"),
        _attr(invoice, "barcode"),
        _value(_raw_attr(order, "quantity")),
        _value(_raw_attr(invoice, "quantity")),
        _value(_quantity_diff(result)),
        _value(_raw_attr(order, "price")),
        _value(_raw_attr(invoice, "price")),
        _value(_raw_attr(order, "amount")),
        _value(_raw_attr(invoice, "amount")),
        _value(_amount_diff(result)),
        _percent_value(_amount_diff_percent(result)),
        _manager_comment(result),
    ]


def _missing_extra_values(result: MatchResult) -> list[object]:
    row = result.order or result.invoice
    order, invoice = result.order, result.invoice
    return [
        result.status,
        _attr(row, "name"),
        _attr(row, "article"),
        _attr(row, "barcode"),
        _value(_raw_attr(order, "quantity")),
        _value(_raw_attr(invoice, "quantity")),
        _value(_raw_attr(order, "price")),
        _value(_raw_attr(invoice, "price")),
        _value(_raw_attr(order, "amount")),
        _value(_raw_attr(invoice, "amount")),
        _manager_comment(result),
    ]


def _manager_comment(result: MatchResult) -> str:
    if result.status == "Нет в счете":
        return "Нет в счёте."
    if result.status == "Лишнее в счете":
        return "Нет в заказе."
    if _has_error(result.order, "Штрихкод пустой"):
        return "В заказе не указан штрихкод."
    if _has_error(result.invoice, "Штрихкод пустой"):
        return "В счёте не указан штрихкод."
    if _has_error(result.order, "Количество не является числом"):
        return "В заказе не прочитано количество."
    if _has_error(result.invoice, "Количество не является числом"):
        return "В счёте не прочитано количество."
    technical = result.technical_comment or result.comment
    if "Проблемная агрегация" in technical or (
        "повторяется" in technical and result.status == "Проверить вручную"
    ):
        return "Позиция повторяется в счёте. Требуется проверка."
    if (
        result.invoice
        and result.invoice.raw.get("aggregation_status") == "success"
        and result.status != "ОК"
    ):
        return (
            "Позиция объединена по повторяющемуся штрихкоду в счёте, "
            "но количество или сумма отличаются."
        )
    if result.status == "Штрихкод не найден":
        return "Нет штрихкода в файле справочника."
    if result.status == "Сумма отличается":
        return "Позиция требует проверки."
    if result.status in {"Проверить вручную", "Ошибка данных"} or (
        "проверить вручную" in result.status.lower()
    ):
        if "Штрихкод не найден" in technical:
            return "Нет штрихкода в файле справочника."
        return "Позиция требует проверки."
    return ""


def _has_error(row: DocumentRow | None, error: str) -> bool:
    return row is not None and error in row.errors


def _name(order: DocumentRow | None, invoice: DocumentRow | None) -> str:
    return _attr(order, "name") or _attr(invoice, "name")


def _attr(row: DocumentRow | None, name: str) -> str:
    return str(getattr(row, name, "") or "") if row else ""


def _raw_attr(row: DocumentRow | None, name: str):
    return getattr(row, name, None) if row else None


def _value(value):
    return float(value) if value is not None else ""


def _percent_value(value):
    return f"{value:.2f}%" if value is not None else ""


def _quantity_diff(result: MatchResult) -> Decimal | None:
    order_qty = _raw_attr(result.order, "quantity")
    invoice_qty = _raw_attr(result.invoice, "quantity")
    if order_qty is None or invoice_qty is None:
        return None
    return invoice_qty - order_qty


def _amount_diff(result: MatchResult) -> Decimal | None:
    order_amount = _raw_attr(result.order, "amount")
    invoice_amount = _raw_attr(result.invoice, "amount")
    if order_amount is None or invoice_amount is None:
        return None
    return invoice_amount - order_amount


def _amount_diff_percent(result: MatchResult) -> Decimal | None:
    order_amount = _raw_attr(result.order, "amount")
    diff = _amount_diff(result)
    if order_amount in {None, 0} or diff is None:
        return None
    try:
        return (diff / order_amount) * Decimal("100")
    except (InvalidOperation, ZeroDivisionError):
        return None


def _show_as_discrepancy(result: MatchResult) -> bool:
    if result.order is None or result.invoice is None:
        return False
    if result.status == "ОК":
        return False
    if result.quantity_status == "Количество отличается":
        return True
    percent = _amount_diff_percent(result)
    if _is_hidden_small_money(result):
        return False
    if _is_manual(result):
        return True
    if percent is None:
        return result.status not in {"Цена отличается", "Проверить вручную"}
    return abs(percent) >= MONEY_DIFF_THRESHOLD_PERCENT


def _hidden_small_money_count(result: CheckResult) -> int:
    return sum(
        1
        for row in result.rows
        if _is_hidden_small_money(row)
    )


def _visible_summary(result: CheckResult) -> dict[str, int]:
    summary = dict(result.summary)
    for row in result.rows:
        if not _is_hidden_small_money(row):
            continue
        if row.status in summary:
            summary[row.status] = max(0, summary[row.status] - 1)
        if _is_manual(row) and "Проверить вручную" in summary:
            summary["Проверить вручную"] = max(0, summary["Проверить вручную"] - 1)
    return summary


def _is_hidden_small_money(result: MatchResult) -> bool:
    if result.order is None or result.invoice is None:
        return False
    if result.quantity_status == "Количество отличается":
        return False
    percent = _amount_diff_percent(result)
    if percent is None or abs(percent) >= MONEY_DIFF_THRESHOLD_PERCENT:
        return False
    technical = result.technical_comment or result.comment
    technical_allowed = (
        not technical
        or technical == "Суммы строк отличаются при совпавших цене и количестве"
        or technical == "Суммы строк отличаются при совпавшем количестве"
    )
    return result.status in {"Цена отличается", "Сумма отличается", "Проверить вручную"} and technical_allowed


def _is_manual(result: MatchResult) -> bool:
    return (
        "проверить вручную" in result.status.lower()
        or result.status in {"Штрихкод не найден", "Ошибка данных"}
    )


def _style(sheet) -> None:
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
