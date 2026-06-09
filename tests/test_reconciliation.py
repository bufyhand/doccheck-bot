from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from core.catalog import Catalog
from core.file_reader import read_document
from core.matcher import reconcile
from core.report_builder import build_report


class ReconciliationTest(TestCase):
    def test_end_to_end_with_shifted_headers_and_different_barcodes(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            order_path = root / "order.xlsx"
            invoice_path = root / "invoice.xlsx"
            report_path = root / "report.xlsx"
            _document(
                order_path,
                [
                    ["Товар A", "001", 2, 10, 20],
                    ["Товар B", "010", 1, 5, 5],
                ],
            )
            _document(
                invoice_path,
                [
                    ["Товар A", "002", 2, 10, 20],
                    ["Товар C", "020", 1, 7, 7],
                ],
            )
            catalog = Catalog(
                items_by_barcode={
                    "001": {"item_id": "a"},
                    "002": {"item_id": "a"},
                    "010": {"item_id": "b"},
                    "020": {"item_id": "c"},
                },
                conflicts={},
            )

            result = reconcile(
                read_document(order_path, "order"),
                read_document(invoice_path, "invoice"),
                catalog,
            )
            report = build_report(result, report_path)

            self.assertEqual(result.summary["ОК"], 1)
            self.assertEqual(result.summary["Нет в счете"], 1)
            self.assertEqual(result.summary["Лишнее в счете"], 1)
            workbook = load_workbook(report, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Итог",
                    "Расхождения",
                    "Нет-Лишнее в счете",
                    "Проверить вручную",
                    "Тех_данные",
                ],
            )
            workbook.close()

    def test_same_unknown_barcode_matches_directly(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            order_path = root / "order.xlsx"
            invoice_path = root / "invoice.xlsx"
            _document(order_path, [["Товар без справочника", "999", 2, 10, 20]])
            _document(invoice_path, [["Товар без справочника", "999", 2, 10, 20]])

            result = reconcile(
                read_document(order_path, "order"),
                read_document(invoice_path, "invoice"),
                Catalog(items_by_barcode={}, conflicts={}),
            )

            self.assertEqual(len(result.rows), 1)
            self.assertEqual(result.rows[0].status, "ОК")
            self.assertIn("сопоставлены напрямую", result.rows[0].comment)

    def test_discounted_invoice_uses_final_amount_as_effective_price(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            invoice_path = root / "invoice.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Артикул",
                    "Товары (работы, услуги)",
                    "Штрихкод",
                    "Количество",
                    "Цена",
                    "Сумма без скидки",
                    "Скидка",
                    "Сумма",
                ]
            )
            sheet.append(
                [
                    "89875",
                    "АКВА МЕНЮ УНИВЕРСАЛ",
                    "4607094650058",
                    1,
                    "66.76",
                    "66.76",
                    "10.01",
                    "56.75",
                ]
            )
            workbook.save(invoice_path)

            invoice = read_document(invoice_path, "invoice")

            self.assertEqual(invoice.columns["amount"], "Сумма")
            self.assertEqual(str(invoice.rows[0].amount), "56.75")
            self.assertEqual(str(invoice.rows[0].discount), "10.01")
            self.assertEqual(str(invoice.rows[0].price), "56.75")

    def test_invoice_with_unit_ean_continuation_row(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            invoice_path = root / "invoice.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([])
            sheet.append([])
            sheet.append(
                [
                    "Но.",
                    None,
                    None,
                    "Описание",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Кол-во",
                    None,
                    None,
                    None,
                    "Ед. Изм.",
                    None,
                    None,
                    None,
                    "Кол-во в Ед. Изм.",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Цена Единицы",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Сумма",
                ]
            )
            sheet.append(
                [
                    "40320008R0",
                    None,
                    None,
                    "Уринари С/О соус 28*85г",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    28,
                    None,
                    None,
                    None,
                    "Штука",
                    None,
                    None,
                    None,
                    1,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    131.028,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    3485.3448,
                ]
            )
            sheet.append([None, None, None, "Unit EAN: 4607134960161   "])
            workbook.save(invoice_path)

            invoice = read_document(invoice_path, "invoice")

            self.assertEqual(len(invoice.rows), 1)
            self.assertEqual(invoice.rows[0].barcode, "4607134960161")
            self.assertEqual(invoice.rows[0].article, "40320008R0")
            self.assertEqual(str(invoice.rows[0].quantity), "28")
            self.assertEqual(str(invoice.rows[0].price), "131.028")
            self.assertEqual(invoice.rows[0].errors, [])

    def test_invoice_with_multiline_vat_table_headers(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            invoice_path = root / "invoice.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([])
            sheet.append(
                [
                    "№ п/п",
                    "Код товара/ работ, услуг",
                    "Наименование товара (описание выполненных работ, оказанных услуг), имущественного права",
                    "Коли-\nчество (объем)",
                    "Цена (тариф) \nза \nединицу \nизмерения",
                    "Стоимость товаров (работ, услуг), имущественных прав без налога - всего",
                    "Сумма налога, предъявляемая покупателю",
                    "Стоимость товаров (работ, услуг), имущественных прав с налогом - всего",
                    "Штрихкод",
                ]
            )
            sheet.append(["служебная подстрока"])
            sheet.append(["А", "Б", "1", "3", "4", "9", None])
            sheet.append(
                [
                    1,
                    "19900521",
                    "PPVD 85г д/к пауч EN Лосось 1/26",
                    6,
                    "108.05",
                    "648.30",
                    "142.62",
                    "790.92",
                    "4600680034416",
                ]
            )
            workbook.save(invoice_path)

            invoice = read_document(invoice_path, "invoice")

            self.assertEqual(len(invoice.rows), 1)
            self.assertEqual(invoice.rows[0].barcode, "4600680034416")
            self.assertEqual(invoice.rows[0].article, "19900521")
            self.assertEqual(invoice.rows[0].quantity, 6)
            self.assertEqual(str(invoice.rows[0].price), "108.05")
            self.assertEqual(str(invoice.rows[0].amount), "790.92")
            self.assertEqual(invoice.rows[0].errors, [])


def _document(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Реквизиты документа"])
    sheet.append(["Еще одна служебная строка"])
    sheet.append(["Наименование", "Штрихкод", "Количество", "Цена", "Сумма"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)
