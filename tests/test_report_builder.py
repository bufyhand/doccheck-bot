from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import load_workbook

from core.models import CheckResult, DocumentRow, MatchResult, ParsedDocument
from core.report_builder import build_report


class ReportBuilderTest(TestCase):
    def test_report_v2_columns_combined_sheet_and_small_money_filter(self):
        with TemporaryDirectory() as directory:
            output = Path(directory) / "report.xlsx"
            result = CheckResult(
                order=ParsedDocument(Path("order.xlsx"), "order", 1, {}, []),
                invoice=ParsedDocument(Path("invoice.xlsx"), "invoice", 1, {}, []),
                rows=[
                    MatchResult(
                        status="Сумма отличается",
                        order=_row("order", "Товар мелкая разница", "111", 10, 100, 1000),
                        invoice=_row("invoice", "Товар мелкая разница", "111", 10, 100, 1003),
                        quantity_status="Количество совпало",
                        price_status="Цена совпала",
                        amount_status="Сумма отличается",
                        comment="Суммы строк отличаются при совпавшем количестве",
                    ),
                    MatchResult(
                        status="Количество отличается",
                        order=_row("order", "Товар количество", "222", 10, 100, 1000),
                        invoice=_row("invoice", "Товар количество", "222", 8, 100, 800),
                        quantity_status="Количество отличается",
                        price_status="Цена совпала",
                        amount_status="Сумма отличается",
                    ),
                    MatchResult(
                        status="Нет в счете",
                        order=_row("order", "Нет", "333", 1, 50, 50),
                    ),
                    MatchResult(
                        status="Лишнее в счете",
                        invoice=_row("invoice", "Лишнее", "444", 2, 70, 140),
                    ),
                ],
            )

            report = build_report(result, output)
            workbook = load_workbook(report, read_only=True, data_only=True)

            sheetnames = workbook.sheetnames

            discrepancies = workbook["Расхождения"]
            headers = [cell.value for cell in next(discrepancies.iter_rows(max_row=1))]
            rows = list(discrepancies.iter_rows(min_row=2, values_only=True))

            missing_extra = workbook["Нет-Лишнее в счете"]
            missing_headers = [
                cell.value for cell in next(missing_extra.iter_rows(max_row=1))
            ]
            statuses = [
                row[0]
                for row in missing_extra.iter_rows(min_row=2, values_only=True)
            ]
            comments = [
                row[-1]
                for row in missing_extra.iter_rows(min_row=2, values_only=True)
            ]

            summary = {
                row[0]: row[1]
                for row in workbook["Итог"].iter_rows(values_only=True)
                if row[0]
            }
            self.assertEqual(
                summary["Скрыто мелких денежных расхождений < 0,5%"], 1
            )
            tech_values = [
                value
                for row in workbook["Тех_данные"].iter_rows(values_only=True)
                for value in row
                if value
            ]
            manual_rows = list(
                workbook["Проверить вручную"].iter_rows(min_row=2, values_only=True)
            )
            workbook.close()

            self.assertEqual(
                sheetnames,
                [
                    "Итог",
                    "Расхождения",
                    "Нет-Лишнее в счете",
                    "Проверить вручную",
                    "Тех_данные",
                ],
            )
            self.assertIn("Разница по количеству", headers)
            self.assertIn("Разница по сумме, ₽", headers)
            self.assertIn("Разница по сумме, %", headers)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0][0], "Количество отличается")
            self.assertEqual(rows[0][8], -2)
            self.assertEqual(rows[0][13], -200)
            self.assertEqual(rows[0][14], "-20.00%")
            self.assertIn(rows[0][15], (None, ""))
            self.assertEqual(missing_headers[0], "Статус")
            self.assertEqual(statuses, ["Нет в счете", "Лишнее в счете"])
            self.assertEqual(comments, ["Нет в счёте.", "Нет в заказе."])
            self.assertIn(
                "Суммы строк отличаются при совпавшем количестве",
                tech_values,
            )
            self.assertEqual(manual_rows, [])


def _row(
    source: str,
    name: str,
    barcode: str,
    quantity: int,
    price: int,
    amount: int,
) -> DocumentRow:
    return DocumentRow(
        source=source,
        row_number=1,
        name=name,
        barcode=barcode,
        quantity=Decimal(quantity),
        price=Decimal(price),
        amount=Decimal(amount),
    )
